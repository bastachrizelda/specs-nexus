import logging
from datetime import datetime, timedelta, timezone
from typing import Dict, Any, Optional
from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy import func, or_, and_, text
from sqlalchemy.orm import Session
from sqlalchemy.exc import ProgrammingError, OperationalError
from pydantic import BaseModel
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart
try:
    from openpyxl.chart.reference import Reference
except ImportError:
    from openpyxl.chart import Reference

from app.database import SessionLocal
from app import models, schemas
from app.auth_utils import get_current_officer

logger = logging.getLogger("app.analytics")

router = APIRouter(prefix="/analytics", tags=["Analytics"])

def get_db():
    db = SessionLocal()
    try:
        # Test database connection
        db.execute(text("SELECT 1"))
        logger.debug("Database connection successful")
        yield db
    except (ProgrammingError, OperationalError) as e:
        logger.error(f"Database error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        logger.error(f"Unexpected error in get_db: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        db.close()

class DateRangeFilter(BaseModel):
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    include_archived: Optional[bool] = False  # New parameter to include archived records

@router.get("/dashboard", response_model=dict)
def get_dashboard_data(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    include_archived: Optional[bool] = False,
    db: Session = Depends(get_db)
) -> Dict[str, Any]:
    logger.debug("Starting dashboard data aggregation")
    
    try:
        # Apply date range filter if provided
        start_date = start_date or (datetime.now(timezone.utc) - timedelta(days=365 * 2))  # Extended to 2 years
        end_date = end_date or datetime.now(timezone.utc)
        include_archived = include_archived or False

        # Validate date range
        if start_date > end_date:
            raise HTTPException(status_code=400, detail="Start date must be before end date")

        # Total BSCS Students (all users in the system)
        try:
            total_cs_students = db.query(models.User).count()
            logger.debug(f"Total BSCS students: {total_cs_students}")
        except ProgrammingError as e:
            logger.error(f"Error querying users table: {str(e)}")
            total_cs_students = 0
            raise HTTPException(status_code=500, detail=f"Error querying users table: {str(e)}")

        no_users = total_cs_students == 0
        if no_users:
            logger.warning("No users found in the users table")

        # Total Specs Members (distinct users with payment_status="Paid" within the date range)
        total_specs_members_query = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date.isnot(None),
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        )
        total_specs_members = total_specs_members_query.distinct().count()
        logger.debug(f"Total Specs members: {total_specs_members}")

        # Semester-specific Specs members
        total_specs_members_first_sem = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.payment_date.isnot(None),
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).distinct().count()
        total_specs_members_second_sem = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.payment_date.isnot(None),
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).distinct().count()
        logger.debug(f"1st Semester Specs members: {total_specs_members_first_sem}, 2nd Semester Specs members: {total_specs_members_second_sem}")

        # None Specs members (users with payment_status="Not Paid" or "Verifying" within the date range)
        none_specs = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status.in_(["Not Paid", "Verifying"]),
            models.Clearance.last_updated.isnot(None),
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).distinct().count()
        none_specs_first_sem = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status.in_(["Not Paid", "Verifying"]),
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.last_updated.isnot(None),
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).distinct().count()
        none_specs_second_sem = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status.in_(["Not Paid", "Verifying"]),
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.last_updated.isnot(None),
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).distinct().count()
        logger.debug(f"None Specs: {none_specs}, 1st Sem: {none_specs_first_sem}, 2nd Sem: {none_specs_second_sem}")

        # Members by requirement (for charts)
        members_by_requirement_raw = db.query(
            models.Clearance.requirement,
            func.count(func.distinct(models.Clearance.user_id))
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date.isnot(None),
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).group_by(models.Clearance.requirement).all()
        members_by_requirement = {req: count for req, count in members_by_requirement_raw}
        logger.debug(f"Members by requirement: {members_by_requirement}")

        # Active members (last 30 days) and recent activity (last 7 days)
        thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
        seven_days_ago = datetime.now(timezone.utc) - timedelta(days=7)
        
        try:
            active_members = db.query(models.User).filter(
                models.User.last_active >= thirty_days_ago,
                models.User.last_active.isnot(None)
            ).count()
            
            recent_activity = db.query(models.User).filter(
                models.User.last_active >= seven_days_ago,
                models.User.last_active.isnot(None)
            ).count()
            
            inactive_members = total_cs_students - active_members
            logger.debug(f"Active members: {active_members}, Inactive members: {inactive_members}, Recent activity (7 days): {recent_activity}")
        except ProgrammingError as e:
            logger.error(f"Error querying active/inactive members: {str(e)}")
            active_members = 0
            inactive_members = total_cs_students
            recent_activity = 0

        # Payment status counts (overall and semester-specific)
        not_paid_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        verifying_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        paid_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()

        not_paid_first_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        not_paid_second_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        verifying_first_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        verifying_second_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        paid_first_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()
        paid_second_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()
        logger.debug(f"Payment Analytics - Not Paid: {not_paid_count}, Verifying: {verifying_count}, Paid: {paid_count}")
        logger.debug(f"1st Sem - Not Paid: {not_paid_first_sem}, Verifying: {verifying_first_sem}, Paid: {paid_first_sem}")
        logger.debug(f"2nd Sem - Not Paid: {not_paid_second_sem}, Verifying: {verifying_second_sem}, Paid: {paid_second_sem}")

        # Payment methods and trends
        payment_methods = db.query(
            models.Clearance.payment_method,
            func.count(models.Clearance.id)
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_method.isnot(None),
            or_(
                and_(
                    models.Clearance.payment_status == "Paid",
                    models.Clearance.payment_date.isnot(None),
                    models.Clearance.payment_date >= start_date,
                    models.Clearance.payment_date <= end_date
                ),
                models.Clearance.payment_status.in_(["Not Paid", "Verifying"])
            )
        ).group_by(models.Clearance.payment_method).all()
        logger.debug(f"Raw payment methods query result: {payment_methods}")
        preferred_payment_methods = [{"method": method, "count": count, "firstSemCount": 0, "secondSemCount": 0} for method, count in payment_methods]

        payment_method_trends = db.query(
            models.Clearance.payment_method,
            models.Clearance.requirement,
            func.count(models.Clearance.id).label('count')
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_method.isnot(None),
            or_(
                and_(
                    models.Clearance.payment_status == "Paid",
                    models.Clearance.payment_date.isnot(None),
                    models.Clearance.payment_date >= start_date,
                    models.Clearance.payment_date <= end_date
                ),
                models.Clearance.payment_status.in_(["Not Paid", "Verifying"])
            )
        ).group_by(models.Clearance.payment_method, models.Clearance.requirement).all()
        logger.debug(f"Raw payment method trends query result: {payment_method_trends}")
        payment_method_trends_dict = {}
        for method, requirement, count in payment_method_trends:
            if method not in payment_method_trends_dict:
                payment_method_trends_dict[method] = {"firstSemCount": 0, "secondSemCount": 0}
            if requirement == "1st Semester Membership":
                payment_method_trends_dict[method]["firstSemCount"] = count
            elif requirement == "2nd Semester Membership":
                payment_method_trends_dict[method]["secondSemCount"] = count
        payment_method_trends_list = [
            {"method": method, "firstSemCount": data["firstSemCount"], "secondSemCount": data["secondSemCount"]}
            for method, data in payment_method_trends_dict.items()
        ]
        for method in preferred_payment_methods:
            for trend in payment_method_trends_list:
                if trend["method"] == method["method"]:
                    method["firstSemCount"] = trend["firstSemCount"]
                    method["secondSemCount"] = trend["secondSemCount"]
        logger.debug(f"Preferred payment methods: {preferred_payment_methods}")
        logger.debug(f"Payment method trends: {payment_method_trends_list}")

        # Payment details by requirement and year
        payment_by_req_year_raw = db.query(
            models.User.year,
            models.Clearance.requirement,
            models.Clearance.payment_status,
            func.count(models.Clearance.id)
        ).join(models.Clearance, models.Clearance.user_id == models.User.id).filter(
            models.Clearance.archived == False,
            or_(
                and_(
                    models.Clearance.payment_status == "Paid",
                    models.Clearance.payment_date >= start_date,
                    models.Clearance.payment_date <= end_date
                ),
                and_(
                    models.Clearance.payment_status.in_(["Not Paid", "Verifying"]),
                    models.Clearance.last_updated >= start_date,
                    models.Clearance.last_updated <= end_date
                )
            )
        ).group_by(models.User.year, models.Clearance.requirement, models.Clearance.payment_status).all()

        byRequirementAndYear = {}
        for user_year, requirement, payment_status, count in payment_by_req_year_raw:
            if not user_year:
                user_year = "Unspecified"
            if requirement not in byRequirementAndYear:
                byRequirementAndYear[requirement] = {}
            if user_year not in byRequirementAndYear[requirement]:
                byRequirementAndYear[requirement][user_year] = {"Not Paid": 0, "Verifying": 0, "Paid": 0}
            byRequirementAndYear[requirement][user_year][payment_status] = count
        logger.debug(f"Payment details by requirement and year: {byRequirementAndYear}")

        # Event details and participation rates
        events_query = db.query(models.Event).filter(
            models.Event.archived == include_archived,  # Use include_archived parameter
            or_(
                models.Event.date >= start_date,
                models.Event.date <= end_date,
                models.Event.date.is_(None)  # Include NULL dates
            )
        )
        events = events_query.all()
        logger.debug(f"Raw events query result: {[(e.title, e.date, e.participant_count, e.archived) for e in events]}")
        events_engagement = []
        events_by_year = {}
        for event in events:
            event_year = event.date.year if event.date else "Unknown"
            engagement = {
                "title": event.title,
                "participant_count": event.participant_count or 0,
                "participation_rate": round((event.participant_count / total_specs_members) * 100, 2) if total_specs_members > 0 and event.participant_count else 0
            }
            events_engagement.append(engagement)
            if event_year not in events_by_year:
                events_by_year[event_year] = []
            events_by_year[event_year].append(engagement)
        popular_events = sorted(events_engagement, key=lambda x: x["participant_count"], reverse=True)[:5]
        logger.debug(f"Event engagement: {events_engagement}")
        logger.debug(f"Popular events: {popular_events}")

        # Clearance by requirement
        clearance_by_requirement_raw = db.query(
            models.Clearance.requirement,
            models.Clearance.status,
            func.count(models.Clearance.id)
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).group_by(models.Clearance.requirement, models.Clearance.status).all()
        clearance_tracking = {}
        for requirement, status, count in clearance_by_requirement_raw:
            if requirement not in clearance_tracking:
                clearance_tracking[requirement] = {"Clear": 0, "Processing": 0, "Not Yet Cleared": 0}
            clearance_tracking[requirement][status] = count
        logger.debug(f"Clearance tracking: {clearance_tracking}")

        # Compliance by year
        compliance_by_year = db.query(
            models.User.year,
            models.Clearance.status,
            func.count(models.Clearance.id)
        ).join(models.Clearance, models.Clearance.user_id == models.User.id).filter(
            models.Clearance.archived == False,
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).group_by(models.User.year, models.Clearance.status).all()

        compliance = {}
        for year, status, count in compliance_by_year:
            year = year or "Unspecified"
            if year not in compliance:
                compliance[year] = {"Clear": 0, "Processing": 0, "Not Yet Cleared": 0}
            compliance[year][status] = count
        logger.debug(f"Compliance by year: {compliance}")

        logger.info("Dashboard data aggregated successfully")
        return {
            "membershipInsights": {
                "totalBSCSStudents": total_cs_students,
                "totalSpecsMembers": total_specs_members,
                "totalSpecsMembersFirstSem": total_specs_members_first_sem,
                "totalSpecsMembersSecondSem": total_specs_members_second_sem,
                "noneSpecs": none_specs,
                "noneSpecsFirstSem": none_specs_first_sem,
                "noneSpecsSecondSem": none_specs_second_sem,
                "activeMembers": active_members,
                "inactiveMembers": inactive_members,
                "recentActivityLast7Days": recent_activity,
                "membersByRequirement": members_by_requirement,
                "noUsers": no_users
            },
            "paymentAnalytics": {
                "byRequirementAndYear": byRequirementAndYear,
                "notPaid": not_paid_count,
                "verifying": verifying_count,
                "paid": paid_count,
                "notPaidFirstSem": not_paid_first_sem,
                "notPaidSecondSem": not_paid_second_sem,
                "verifyingFirstSem": verifying_first_sem,
                "verifyingSecondSem": verifying_second_sem,
                "paidFirstSem": paid_first_sem,
                "paidSecondSem": paid_second_sem,
                "preferredPaymentMethods": preferred_payment_methods,
                "paymentMethodTrends": payment_method_trends_list
            },
            "eventsEngagement": {
                "events": events_engagement,
                "popularEvents": popular_events,
                "breakdownByYear": events_by_year
            },
            "clearanceTracking": {
                "byRequirement": clearance_tracking,
                "complianceByYear": compliance
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error aggregating dashboard data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Internal server error while aggregating dashboard data: {str(e)}")

@router.get("/report/officer-dashboard")
def generate_officer_dashboard_report(
    db: Session = Depends(get_db),
    current_officer: models.Officer = Depends(get_current_officer)
):
    """Generate Excel report for officer dashboard with charts"""
    logger.debug(f"Officer {current_officer.id} generating dashboard report")
    
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Get dashboard data by calling the endpoint logic directly
        start_date = datetime.now(timezone.utc) - timedelta(days=365 * 2)
        end_date = datetime.now(timezone.utc)
        include_archived = False
        
        # Reuse the same logic from get_dashboard_data
        total_cs_students = db.query(models.User).count()
        total_specs_members = db.query(models.Clearance.user_id).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date.isnot(None),
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).distinct().count()
        
        thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
        active_members = db.query(models.User).filter(
            models.User.last_active >= thirty_days_ago,
            models.User.last_active.isnot(None)
        ).count()
        inactive_members = total_cs_students - active_members
        
        # Payment Analytics - All Semesters
        verifying_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        paid_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()
        not_paid_count = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        
        # Payment Analytics - 1st Semester
        not_paid_first_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        verifying_first_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        paid_first_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "1st Semester Membership",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()
        
        # Payment Analytics - 2nd Semester
        not_paid_second_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Not Paid",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        verifying_second_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Verifying",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).count()
        paid_second_sem = db.query(models.Clearance).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_status == "Paid",
            models.Clearance.requirement == "2nd Semester Membership",
            models.Clearance.payment_date >= start_date,
            models.Clearance.payment_date <= end_date
        ).count()
        
        # Payment methods
        payment_analytics_data = db.query(
            models.Clearance.payment_method,
            func.count(models.Clearance.id)
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.payment_method.isnot(None),
            or_(
                and_(
                    models.Clearance.payment_status == "Paid",
                    models.Clearance.payment_date.isnot(None),
                    models.Clearance.payment_date >= start_date,
                    models.Clearance.payment_date <= end_date
                ),
                models.Clearance.payment_status.in_(["Not Paid", "Verifying"])
            )
        ).group_by(models.Clearance.payment_method).all()
        
        # Payment details by requirement and year
        payment_by_req_year_raw = db.query(
            models.User.year,
            models.Clearance.requirement,
            models.Clearance.payment_status,
            func.count(models.Clearance.id)
        ).join(models.Clearance, models.Clearance.user_id == models.User.id).filter(
            models.Clearance.archived == False,
            or_(
                and_(
                    models.Clearance.payment_status == "Paid",
                    models.Clearance.payment_date >= start_date,
                    models.Clearance.payment_date <= end_date
                ),
                and_(
                    models.Clearance.payment_status.in_(["Not Paid", "Verifying"]),
                    models.Clearance.last_updated >= start_date,
                    models.Clearance.last_updated <= end_date
                )
            )
        ).group_by(models.User.year, models.Clearance.requirement, models.Clearance.payment_status).all()
        
        byRequirementAndYear = {}
        for user_year, requirement, payment_status, count in payment_by_req_year_raw:
            if not user_year:
                user_year = "Unspecified"
            if requirement not in byRequirementAndYear:
                byRequirementAndYear[requirement] = {}
            if user_year not in byRequirementAndYear[requirement]:
                byRequirementAndYear[requirement][user_year] = {"Not Paid": 0, "Verifying": 0, "Paid": 0}
            byRequirementAndYear[requirement][user_year][payment_status] = count
        
        # Events data
        events_query = db.query(models.Event).filter(
            models.Event.archived == include_archived,
            or_(
                models.Event.date >= start_date,
                models.Event.date <= end_date,
                models.Event.date.is_(None)
            )
        )
        events_data = events_query.all()
        
        events_engagement = []
        for event in events_data:
            engagement = {
                "title": event.title,
                "participant_count": event.participant_count or 0,
                "participation_rate": round((event.participant_count / total_specs_members) * 100, 2) if total_specs_members > 0 and event.participant_count else 0
            }
            events_engagement.append(engagement)
        
        popular_events = sorted(events_engagement, key=lambda x: x["participant_count"], reverse=True)[:5]
        
        # Clearance tracking
        clearance_data = db.query(
            models.Clearance.requirement,
            models.Clearance.status,
            func.count(models.Clearance.id)
        ).filter(
            models.Clearance.archived == False,
            models.Clearance.last_updated >= start_date,
            models.Clearance.last_updated <= end_date
        ).group_by(models.Clearance.requirement, models.Clearance.status).all()
        
        clearance_tracking = {}
        for requirement, status, count in clearance_data:
            if requirement not in clearance_tracking:
                clearance_tracking[requirement] = {"Clear": 0, "Processing": 0, "Not Yet Cleared": 0}
            clearance_tracking[requirement][status] = count
        
        # Sheet 1: Student Insights
        ws_students = wb.create_sheet("Student Insights")
        ws_students.append(["Student Insights Report"])
        ws_students.append([])
        ws_students.append(["Total BSCS Students", total_cs_students])
        ws_students.append(["Active Members", active_members])
        ws_students.append(["Inactive Members", inactive_members])
        ws_students.append([])
        
        # Get all users
        users = db.query(models.User).all()
        ws_students.append(["ID", "Name", "Student Number", "Year", "Block", "Email", "Status"])
        for user in users:
            is_active = False
            if user.last_active:
                # Convert naive datetime to UTC-aware if needed
                last_active = user.last_active
                if last_active.tzinfo is None:
                    last_active = last_active.replace(tzinfo=timezone.utc)
                is_active = last_active >= thirty_days_ago
            ws_students.append([
                user.id,
                user.full_name or "N/A",
                user.student_number or "N/A",
                user.year or "N/A",
                user.block or "N/A",
                user.email or "N/A",
                "Active" if is_active else "Inactive"
            ])
        
        # Sheet 2: Payment Analytics
        ws_payments = wb.create_sheet("Payment Analytics")
        ws_payments.append(["Payment Analytics Report"])
        ws_payments.append([])
        
        # Payment Analytics - All Semesters (data and chart)
        ws_payments.append(["Payment Status - All Semesters"])
        ws_payments.append(["Status", "Count"])
        all_sem_data_start = ws_payments.max_row + 1
        ws_payments.append(["Not Paid", not_paid_count])
        ws_payments.append(["Verifying", verifying_count])
        ws_payments.append(["Paid", paid_count])
        all_sem_data_end = ws_payments.max_row
        ws_payments.append([])
        
        # Chart 1: Payment Analytics - All Semesters (Pie Chart)
        if all_sem_data_start <= all_sem_data_end:
            try:
                chart_all_sem = PieChart()
                chart_all_sem.title = "Payment Analytics - All Semesters"
                chart_all_sem.width = 10
                chart_all_sem.height = 7
                data_all = Reference(ws_payments, min_col=2, min_row=all_sem_data_start, max_row=all_sem_data_end)
                labels_all = Reference(ws_payments, min_col=1, min_row=all_sem_data_start, max_row=all_sem_data_end)
                chart_all_sem.add_data(data_all, titles_from_data=False)
                chart_all_sem.set_categories(labels_all)
                ws_payments.add_chart(chart_all_sem, "A8")
            except Exception as chart_error:
                logger.warning(f"Failed to create all semesters chart: {str(chart_error)}")
        
        # Payment Analytics - 1st Semester (data and chart)
        ws_payments.append([])
        ws_payments.append(["Payment Status - 1st Semester"])
        ws_payments.append(["Status", "Count"])
        first_sem_data_start = ws_payments.max_row + 1
        ws_payments.append(["Not Paid", not_paid_first_sem])
        ws_payments.append(["Verifying", verifying_first_sem])
        ws_payments.append(["Paid", paid_first_sem])
        first_sem_data_end = ws_payments.max_row
        ws_payments.append([])
        
        # Chart 2: Payment Analytics - 1st Semester (Pie Chart)
        if first_sem_data_start <= first_sem_data_end:
            try:
                chart_first_sem = PieChart()
                chart_first_sem.title = "Payment Analytics - 1st Semester"
                chart_first_sem.width = 10
                chart_first_sem.height = 7
                data_first = Reference(ws_payments, min_col=2, min_row=first_sem_data_start, max_row=first_sem_data_end)
                labels_first = Reference(ws_payments, min_col=1, min_row=first_sem_data_start, max_row=first_sem_data_end)
                chart_first_sem.add_data(data_first, titles_from_data=False)
                chart_first_sem.set_categories(labels_first)
                ws_payments.add_chart(chart_first_sem, "D8")
            except Exception as chart_error:
                logger.warning(f"Failed to create 1st semester chart: {str(chart_error)}")
        
        # Payment Analytics - 2nd Semester (data and chart)
        ws_payments.append([])
        ws_payments.append(["Payment Status - 2nd Semester"])
        ws_payments.append(["Status", "Count"])
        second_sem_data_start = ws_payments.max_row + 1
        ws_payments.append(["Not Paid", not_paid_second_sem])
        ws_payments.append(["Verifying", verifying_second_sem])
        ws_payments.append(["Paid", paid_second_sem])
        second_sem_data_end = ws_payments.max_row
        ws_payments.append([])
        
        # Chart 3: Payment Analytics - 2nd Semester (Pie Chart)
        if second_sem_data_start <= second_sem_data_end:
            try:
                chart_second_sem = PieChart()
                chart_second_sem.title = "Payment Analytics - 2nd Semester"
                chart_second_sem.width = 10
                chart_second_sem.height = 7
                data_second = Reference(ws_payments, min_col=2, min_row=second_sem_data_start, max_row=second_sem_data_end)
                labels_second = Reference(ws_payments, min_col=1, min_row=second_sem_data_start, max_row=second_sem_data_end)
                chart_second_sem.add_data(data_second, titles_from_data=False)
                chart_second_sem.set_categories(labels_second)
                ws_payments.add_chart(chart_second_sem, "G8")
            except Exception as chart_error:
                logger.warning(f"Failed to create 2nd semester chart: {str(chart_error)}")
        
        # Payment Methods
        ws_payments.append([])
        ws_payments.append(["Payment Method", "Count"])
        payment_method_start_row = ws_payments.max_row + 1
        for method, count in payment_analytics_data:
            ws_payments.append([method or "N/A", count])
        
        # Sheet 3: Payment Details by Requirement & Year
        ws_payment_details = wb.create_sheet("Payment Details by Req & Year")
        ws_payment_details.append(["Payment Details by Requirement & Year"])
        ws_payment_details.append([])
        
        # Prepare data for chart
        row_num = 3
        ws_payment_details.append(["Year", "Requirement", "Not Paid", "Verifying", "Paid"])
        chart_data_start_row = row_num
        
        for requirement, year_data in byRequirementAndYear.items():
            for year, status_data in year_data.items():
                ws_payment_details.append([
                    year,
                    requirement,
                    status_data.get("Not Paid", 0),
                    status_data.get("Verifying", 0),
                    status_data.get("Paid", 0)
                ])
                row_num += 1
        
        # Chart 4: Payment Details by Requirement & Year (Bar Chart)
        if row_num > chart_data_start_row:
            try:
                chart_payment_details = BarChart()
                chart_payment_details.type = "col"
                chart_payment_details.style = 10
                chart_payment_details.title = "Payment Details by Requirement & Year"
                chart_payment_details.y_axis.title = "Count"
                chart_payment_details.x_axis.title = "Year & Requirement"
                chart_payment_details.width = 15
                chart_payment_details.height = 10
                
                data_payment = Reference(ws_payment_details, min_col=3, min_row=chart_data_start_row+1, max_col=5, max_row=row_num)
                cats_payment = Reference(ws_payment_details, min_col=1, min_row=chart_data_start_row+1, max_row=row_num)
                chart_payment_details.add_data(data_payment, titles_from_data=True)
                chart_payment_details.set_categories(cats_payment)
                ws_payment_details.add_chart(chart_payment_details, "A" + str(row_num + 3))
            except Exception as chart_error:
                logger.warning(f"Failed to create payment details chart: {str(chart_error)}")
        
        # Sheet 4: Events Engagement
        ws_events = wb.create_sheet("Events Engagement")
        ws_events.append(["Events Engagement Report"])
        ws_events.append([])
        ws_events.append(["Event Title", "Participant Count", "Participation Rate (%)"])
        events_start_row = 4
        for engagement in events_engagement:
            ws_events.append([
                engagement["title"],
                engagement["participant_count"],
                engagement["participation_rate"]
            ])
        
        # Chart 5: Events Engagement (Bar Chart)
        if ws_events.max_row > events_start_row:
            try:
                chart_events = BarChart()
                chart_events.type = "col"
                chart_events.style = 10
                chart_events.title = "Events Engagement"
                chart_events.y_axis.title = "Participant Count"
                chart_events.x_axis.title = "Events"
                chart_events.width = 15
                chart_events.height = 10
                
                data_events = Reference(ws_events, min_col=2, min_row=events_start_row-1, max_row=ws_events.max_row)
                cats_events = Reference(ws_events, min_col=1, min_row=events_start_row, max_row=ws_events.max_row)
                chart_events.add_data(data_events, titles_from_data=True)
                chart_events.set_categories(cats_events)
                ws_events.add_chart(chart_events, "E3")
            except Exception as chart_error:
                logger.warning(f"Failed to create events engagement chart: {str(chart_error)}")
        
        # Sheet 5: Popular Events
        ws_popular = wb.create_sheet("Popular Events")
        ws_popular.append(["Top 5 Popular Events"])
        ws_popular.append([])
        ws_popular.append(["Event Title", "Participant Count", "Participation Rate (%)"])
        popular_start_row = 4
        for event in popular_events:
            ws_popular.append([
                event["title"],
                event["participant_count"],
                event["participation_rate"]
            ])
        
        # Chart 6: Popular Events (Bar Chart)
        if ws_popular.max_row > popular_start_row:
            try:
                chart_popular = BarChart()
                chart_popular.type = "col"
                chart_popular.style = 10
                chart_popular.title = "Top 5 Popular Events"
                chart_popular.y_axis.title = "Participant Count"
                chart_popular.x_axis.title = "Events"
                chart_popular.width = 12
                chart_popular.height = 8
                
                data_popular = Reference(ws_popular, min_col=2, min_row=popular_start_row-1, max_row=ws_popular.max_row)
                cats_popular = Reference(ws_popular, min_col=1, min_row=popular_start_row, max_row=ws_popular.max_row)
                chart_popular.add_data(data_popular, titles_from_data=True)
                chart_popular.set_categories(cats_popular)
                ws_popular.add_chart(chart_popular, "E3")
            except Exception as chart_error:
                logger.warning(f"Failed to create popular events chart: {str(chart_error)}")
        
        # Sheet 6: Clearance Tracking
        ws_clearance = wb.create_sheet("Clearance Tracking")
        ws_clearance.append(["Clearance Tracking Report"])
        ws_clearance.append([])
        ws_clearance.append(["Requirement", "Clear", "Processing", "Not Yet Cleared"])
        clearance_start_row = 4
        
        # Prepare data for stacked bar chart
        requirements_list = []
        clear_counts = []
        processing_counts = []
        not_cleared_counts = []
        
        for requirement, status_data in clearance_tracking.items():
            requirements_list.append(requirement or "N/A")
            clear_counts.append(status_data.get("Clear", 0))
            processing_counts.append(status_data.get("Processing", 0))
            not_cleared_counts.append(status_data.get("Not Yet Cleared", 0))
            ws_clearance.append([
                requirement or "N/A",
                status_data.get("Clear", 0),
                status_data.get("Processing", 0),
                status_data.get("Not Yet Cleared", 0)
            ])
        
        # Chart 7: Clearance Tracking (Stacked Bar Chart)
        if ws_clearance.max_row > clearance_start_row:
            try:
                chart_clearance = BarChart()
                chart_clearance.type = "col"
                chart_clearance.style = 10
                chart_clearance.title = "Clearance Tracking by Requirement"
                chart_clearance.y_axis.title = "Count"
                chart_clearance.x_axis.title = "Requirement"
                chart_clearance.grouping = "stacked"
                chart_clearance.width = 15
                chart_clearance.height = 10
                
                data_clearance = Reference(ws_clearance, min_col=2, min_row=clearance_start_row-1, max_col=4, max_row=ws_clearance.max_row)
                cats_clearance = Reference(ws_clearance, min_col=1, min_row=clearance_start_row, max_row=ws_clearance.max_row)
                chart_clearance.add_data(data_clearance, titles_from_data=True)
                chart_clearance.set_categories(cats_clearance)
                ws_clearance.add_chart(chart_clearance, "A" + str(ws_clearance.max_row + 3))
            except Exception as chart_error:
                logger.warning(f"Failed to create clearance tracking chart: {str(chart_error)}")
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Style data headers (row 3 or 4 depending on sheet)
            header_row = 3
            if ws.title == "Events Engagement" or ws.title == "Popular Events" or ws.title == "Clearance Tracking":
                header_row = 4
            elif ws.title == "Payment Details by Req & Year":
                header_row = 3
            
            for cell in ws[header_row]:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.font = Font(bold=True)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Officer {current_officer.id} generated dashboard report with charts successfully")
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=officer-dashboard-report-{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.xlsx"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        logger.error(f"Error generating report: {str(e)}\n{error_traceback}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")